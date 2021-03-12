
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import os.path

import boto3
from boto3.s3.transfer import S3Transfer
client=boto3.client('s3',aws_access_key_id='AKIAI36R2ZXCGW4NOOMQ',aws_secret_access_key='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')
transfer=S3Transfer(client)

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
from openpyxl.styles import Font



global text1,text2,cb

if(os.path.exists('data.xlsx')):
                                                        
                                    window=Tk()
                                    window.title('PLASMA 3.0')
                                    window.geometry('1920x1080')
                                    window.configure(bg='skyblue')
                                    
                                        
                                    
                                    def click3():
                                        
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m,column=8).value=select.get()
                                        sheet.cell(row=m,column=8).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        messagebox.showinfo('PLASMA 3.0',"     EVENTS WITH ELEGANCE\n    THANK YOU FOR REGISTERING")
                                        window3.destroy()
                                    
                                    def click2():
                                        
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m,column=6).value=textw1.get()
                                        sheet.cell(row=m,column=6).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m,column=7).value=textw21.get()
                                        sheet.cell(row=m,column=7).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        if(sheet.cell(row=m,column=4).value==sheet.cell(row=m,column=6).value and sheet.cell(row=m,column=5).value==sheet.cell(row=m,column=7).value):
                                                    transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                                    window2.destroy()
                                                    global window3
                                                    window3=Tk()
                                                    window3.title('LOGIN NOW...')
                                                    window3.geometry('1920x1080')
                                                    window3.configure(bg='SKY BLUE')
                                                    lb=Label(window3,text='Welcome...\nSelect the events to participate',bg='SKY BLUE',fg='RED',font=('OPTIMUS',30))
                                                    lb.pack()
                                                    
                                                    
                                                    lb=Label(window3,text='TECHNICAL EVENTS',bg='SKY BLUE',fg='YELLOW',font=('OPTIMUS',25))
                                                    lb.place(x=250,y=200)
                                                    
                                                    lb=Label(window3,text='NON-TECHNICAL EVENTS',bg='SKY BLUE',fg='YELLOW',font=('OPTIMUS',25))
                                                    lb.place(x=850,y=200)
                                                    
                                                    global select
                                                    
                                                    select=StringVar()
                                                
                                                    b=Radiobutton(window3,text='ROBO RACE',value='ROBO RACE',bg='skyblue',variable=select)
                                                    b.place(x=350,y=300)
                                                    
                                                    b1=Radiobutton(window3,text='ROBO SOCCER',value='ROBO SOCCER',bg='skyblue',variable=select)
                                                    b1.place(x=350,y=350)
                                                    
                                                    b2=Radiobutton(window3,text='LINE FOLLOWER',value='LINE FOLLOWER',bg='skyblue',variable=select)
                                                    b2.place(x=350,y=400)
                                                    
                                                    b3=Radiobutton(window3,text='CIRCUIT DEBUGGING',value='CIRCUIT DEBUGGING',bg='skyblue',variable=select)
                                                    b3.place(x=950,y=300)
                                                    
                                                    b4=Radiobutton(window3,text='TEST C',value='TEST C',bg='skyblue',variable=select)
                                                    b4.place(x=950,y=350)
                                                    
                                                    b5=Radiobutton(window3,text='GET SET GO',value='GET SET GO',bg='skyblue',variable=select)
                                                    b5.place(x=350,y=450)
                                                    
                                                    b6=Radiobutton(window3,text='PAPER PRESENTATION',value='PAPER PRESENTATION',bg='skyblue',variable=select)
                                                    b6.place(x=950,y=400)
                                                     
                                                    
                                                    button=Button(window3,text='SUBMIT',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click3)
                                                    button.place(x=700,y=550)
                                        else:
                                                        messagebox.showinfo('PLASMA 3.0'," Enter Valid User Id & Password ")
                                        
                                    
                                    def click1():
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m,column=4).value=textw.get()
                                        sheet.cell(row=m,column=4).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m,column=5).value=textw2.get()
                                        sheet.cell(row=m,column=5).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                    
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        window1.destroy()
                                        global window2
                                        window2=Tk()
                                        window2.title('LOGIN NOW...')
                                        window2.geometry('1920x1080')
                                        window2.configure(bg='light green')
                                        
                                        global textw1,textw21
                                        
                                        lb=Label(window2,text='ENTER USER NAME AND PASSWORD',bg='light green',fg='darkblue',font=('tesla',30))
                                        lb.pack()
                                        
                                        lbw1=Label(window2,text='USER NAME --',bg='light green',fg='red',font=('optimus',20))
                                        lbw1.place(x=300,y=100)
                                        textw1=Entry(window2,width=30,fg='purple',font=('OCR A std',20))
                                        textw1.place(x=520,y=100)
                                        
                                        lbw21=Label(window2,text='PASSWORD --',bg='light green',fg='red',font=('optimus',20))
                                        lbw21.place(x=300,y=200)
                                        textw21=Entry(window2,width=30,fg='purple',font=('OCR A std',20))
                                        textw21.place(x=520,y=200)
                                        
                                        button=Button(window2,text='login',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click2)
                                        button.place(x=750,y=250)
                                        
                                        
                                                
                                        
                                            
                                            
                                        
                                        
                                    def click():
                                        global m
                                        
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m+1,column=1).value=text1.get()
                                        sheet.cell(row=m+1,column=1).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m+1,column=2).value=text2.get()
                                        sheet.cell(row=m+1,column=2).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m+1,column=3).value=cb.get()
                                        sheet.cell(row=m+1,column=3).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        window.destroy()
                                        global window1
                                        window1=Tk()
                                        window1.title('Signup here...')
                                        window1.geometry('1920x1080')
                                        window1.configure(bg='pink')
                                        
                                        global textw,textw2
                                        
                                        
                                        lb=Label(window1,text='CREATE UR ACCOUNT',bg='PINK',fg='red',font=('optimus',30))
                                        lb.pack()
                                        
                                        lbw=Label(window1,text='USER NAME --',bg='pink',fg='red',font=('optimus',20))
                                        lbw.place(x=300,y=100)
                                        textw=Entry(window1,width=30,fg='purple',font=('OCR A std',20))
                                        textw.place(x=520,y=100)
                                        
                                        lbw2=Label(window1,text='PASSWORD --',bg='pink',fg='red',font=('optimus',20))
                                        lbw2.place(x=300,y=200)
                                        textw2=Entry(window1,width=30,fg='purple',font=('OCR A std',20))
                                        textw2.place(x=520,y=200)
                                        
                                        button=Button(window1,text='SIGN UP',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click1)
                                        button.place(x=700,y=250)
                                        
                                        button=Button(window1,text='login',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click1)
                                        button.place(x=850,y=250)
                                    
                                
                                    
                                    
                                    lb=Label(window,text='WELCOME TO JNNCE',bg='skyblue',fg='red',font=('samarkan',30))
                                    lb.pack()
                                    
                                    lb=Label(window,text='PLASMA 3.0',bg='skyblue',fg='red',font=('optimus',30))
                                    lb.pack()
                                    
                                    lb=Label(window,text='ENTER DETAILS TO REGISTER FOR EVENTS',bg='skyblue',fg='red',font=('optimus',30))
                                    lb.pack()
                                    
                                    global text1,text2,cb
                                    
                                    lb1=Label(window,text='NAME      --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb1.place(x=350,y=300)
                                    text1=Entry(window,width=30,fg='purple',font=('OCR A std',20))
                                    text1.place(x=550,y=300)
                                    
                                    lb2=Label(window,text='COLLEGE   --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb2.place(x=350,y=400)
                                    text2=Entry(window,width=30,fg='purple',font=('OCR A std',20))
                                    text2.place(x=550,y=400)
                                    
                                    lb3=Label(window,text='SEM        --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb3.place(x=350,y=500)
                                    
                                    lb4=Label(window,text='(*Already have an account login)',bg='skyblue',fg='red',font=('optimus',10))
                                    lb4.place(x=820,y=620)

                                    
                                    
                                    br=['1','2','3','4','5','6','7','8']
                                    cb=ttk.Combobox(window,values=br,width=92)
                                    cb.place(x=550,y=505)
                                    
                                    button=Button(window,text='NEXT',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click)
                                    button.place(x=800,y=570)
                                    
                                   
                                   
else:
                        
                                    book=Workbook()
                                    sheet=book.active
                                    sheet.column_dimensions['A'].width = 20
                                    sheet.column_dimensions['B'].width = 20
                                    sheet.column_dimensions['C'].width = 20
                                    sheet.column_dimensions['D'].width = 20
                                    sheet.column_dimensions['E'].width = 20
                                    sheet.column_dimensions['F'].width = 20
                                    sheet.column_dimensions['G'].width = 20
                                    sheet.column_dimensions['H'].width = 20
                                    
                                    sheet['A1']='NAME'
                                    sheet['A1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['A1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['B1']='COLLEGE'
                                    sheet['B1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['B1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['C1']='SEM'
                                    sheet['C1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['C1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['D1']='USER NAME'
                                    sheet['D1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['D1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['E1']='PASSWORD'
                                    sheet['E1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['E1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['F1']='USER NAME'
                                    sheet['F1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['F1'].alignment=Alignment(horizontal='center')
                                    
                                    sheet['G1']='PASSWORD'
                                    sheet['G1'].font=Font(bold=True,size=9,color='FF0000FF',name='Futura Md BT')
                                    sheet['G1'].alignment=Alignment(horizontal='center')
                                    
                                    
                                    sheet['H1']='EVENTS'
                                    sheet['H1'].font=Font(bold=True,size=9,color='FF0000FF',name='FURTHER')
                                    sheet['H1'].alignment=Alignment(horizontal='center')
                                    
                                    book.save('data.xlsx')
                    
                    
                                               
                                    window=Tk()
                                    window.title('PLASMA 3.0')
                                    window.geometry('1920x1080')
                                    window.configure(bg='skyblue')
                                    
                                    def click3():
                                        
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m,column=8).value=select.get()
                                        sheet.cell(row=m,column=8).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        messagebox.showinfo('PLASMA 3.0',"     EVENTS WITH ELEGANCE\n             THANK YOU")
                                        window3.destroy()
                                    
                                    def click2():
                                                book=load_workbook('data.xlsx')
                                                sheet=book.active
                                                m=sheet.max_row
                                                sheet.cell(row=m,column=6).value=textw1.get()
                                                sheet.cell(row=m,column=6).alignment=Alignment(horizontal='center')
                                                sheet.cell(row=m,column=7).value=textw21.get()
                                                sheet.cell(row=m,column=7).alignment=Alignment(horizontal='center')
                                                book.save('data.xlsx')
                                                if(sheet.cell(row=m,column=4).value==sheet.cell(row=m,column=6).value and sheet.cell(row=m,column=5).value==sheet.cell(row=m,column=7).value):
                                                
                                                    transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                                    window2.destroy()
                                                    global window3
                                                    window3=Tk()
                                                    window3.title('LOGIN NOW...')
                                                    window3.geometry('1920x1080')
                                                    window3.configure(bg='SKY BLUE')
                                                    lb=Label(window3,text='Welcome...\nSelect the events to participate',bg='SKY BLUE',fg='RED',font=('OPTIMUS',30))
                                                    lb.pack()
                                                    
                                                    
                                                    lb=Label(window3,text='TECHNICAL EVENTS',bg='SKY BLUE',fg='YELLOW',font=('OPTIMUS',25))
                                                    lb.place(x=250,y=200)
                                                    
                                                    lb=Label(window3,text='NON-TECHNICAL EVENTS',bg='SKY BLUE',fg='YELLOW',font=('OPTIMUS',25))
                                                    lb.place(x=850,y=200)
                                                    
                                                    global select
                                                    
                                                    select=StringVar()
                                                
                                                    b=Radiobutton(window3,text='ROBO RACE',value='ROBO RACE',bg='skyblue',variable=select)
                                                    b.place(x=350,y=300)
                                                    
                                                    b1=Radiobutton(window3,text='ROBO SOCCER',value='ROBO SOCCER',bg='skyblue',variable=select)
                                                    b1.place(x=350,y=350)
                                                    
                                                    b2=Radiobutton(window3,text='LINE FOLLOWER',value='LINE FOLLOWER',bg='skyblue',variable=select)
                                                    b2.place(x=350,y=400)
                                                    
                                                    b3=Radiobutton(window3,text='CIRCUIT DEBUGGING',value='CIRCUIT DEBUGGING',bg='skyblue',variable=select)
                                                    b3.place(x=950,y=300)
                                                    
                                                    b4=Radiobutton(window3,text='TEST C',value='TEST C',bg='skyblue',variable=select)
                                                    b4.place(x=950,y=350)
                                                    
                                                    b5=Radiobutton(window3,text='GET SET GO',value='GET SET GO',bg='skyblue',variable=select)
                                                    b5.place(x=350,y=450)
                                                    
                                                    b6=Radiobutton(window3,text='PAPER PRESENTATION',value='PAPER PRESENTATION',bg='skyblue',variable=select)
                                                    b6.place(x=950,y=400)
                                                     
                                                    
                                                    button=Button(window3,text='SUBMIT',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click3)
                                                    button.place(x=700,y=550)
                                                else:
                                                        messagebox.showinfo('PLASMA 3.0'," Enter Valid User Id & Password")
                                            
                                        
                                    
                                    def click1():
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m,column=4).value=textw.get()
                                        sheet.cell(row=m,column=4).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m,column=5).value=textw2.get()
                                        sheet.cell(row=m,column=5).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        window1.destroy()
                                        global window2
                                        window2=Tk()
                                        window2.title('LOGIN NOW...')
                                        window2.geometry('1920x1080')
                                        window2.configure(bg='light green')
                                        
                                        global textw1,textw21
                                        
                                        lb=Label(window2,text='ENTER USER NAME AND PASSWORD',bg='light green',fg='darkblue',font=('tesla',30))
                                        lb.pack()
                                        
                                        lbw1=Label(window2,text='USER NAME --',bg='light green',fg='red',font=('optimus',20))
                                        lbw1.place(x=300,y=100)
                                        textw1=Entry(window2,width=30,fg='purple',font=('OCR A std',20))
                                        textw1.place(x=520,y=100)
                                        
                                        lbw21=Label(window2,text='PASSWORD --',bg='light green',fg='red',font=('optimus',20))
                                        lbw21.place(x=300,y=200)
                                        textw21=Entry(window2,width=30,fg='purple',font=('OCR A std',20))
                                        textw21.place(x=520,y=200)
                                        
                                        button=Button(window2,text='login',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click2)
                                        button.place(x=750,y=250)
                                        
                            
                                    
                                        
                                        
                                        
                                        
                                    
                                    
                                    
                                    def click():
                                        global m
                                        
                                        book=load_workbook('data.xlsx')
                                        sheet=book.active
                                        m=sheet.max_row
                                        sheet.cell(row=m+1,column=1).value=text1.get()
                                        sheet.cell(row=m+1,column=1).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m+1,column=2).value=text2.get()
                                        sheet.cell(row=m+1,column=2).alignment=Alignment(horizontal='center')
                                        sheet.cell(row=m+1,column=3).value=cb.get()
                                        sheet.cell(row=m+1,column=3).alignment=Alignment(horizontal='center')
                                        book.save('data.xlsx')
                                        transfer.upload_file(r'data.xlsx','travelheaven4',r'data.xlsx')
                                        window.destroy()
                                        global window1
                                        window1=Tk()
                                        window1.title('Signup here...')
                                        window1.geometry('1920x1080')
                                        window1.configure(bg='pink')
                                        
                                        global textw,textw2
                                        
                                        
                                        lb=Label(window1,text='CREATE UR ACCOUNT',bg='PINK',fg='red',font=('optimus',30))
                                        lb.pack()
                                        
                                        lbw=Label(window1,text='USER NAME --',bg='pink',fg='red',font=('optimus',20))
                                        lbw.place(x=300,y=100)
                                        textw=Entry(window1,width=30,fg='purple',font=('OCR A std',20))
                                        textw.place(x=520,y=100)
                                        
                                        lbw2=Label(window1,text='PASSWORD --',bg='pink',fg='red',font=('optimus',20))
                                        lbw2.place(x=300,y=200)
                                        textw2=Entry(window1,width=30,fg='purple',font=('OCR A std',20))
                                        textw2.place(x=520,y=200)
                                        
                                        button=Button(window1,text='SIGN UP',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click1)
                                        button.place(x=700,y=250)
                                        
                                        button=Button(window1,text='login',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click1)
                                        button.place(x=850,y=250)
                                    
                                    
                                    
                                    
                                    
                                    lb=Label(window,text='WELCOME TO JNNCE',bg='skyblue',fg='red',font=('samarkan',30))
                                    lb.pack()
                                    
                                    lb=Label(window,text='PLASMA 3.0',bg='skyblue',fg='red',font=('optimus',30))
                                    lb.pack()
                                    
                                    lb=Label(window,text='ENTER DETAILS TO REGISTER FOR EVENTS',bg='skyblue',fg='red',font=('optimus',30))
                                    lb.pack()
                                    
                                    
                                    
                                    lb1=Label(window,text='NAME      --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb1.place(x=350,y=300)
                                    text1=Entry(window,width=30,fg='purple',font=('OCR A std',20))
                                    text1.place(x=550,y=300)
                                    
                                    lb2=Label(window,text='COLLEGE   --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb2.place(x=350,y=400)
                                    text2=Entry(window,width=30,fg='purple',font=('OCR A std',20))
                                    text2.place(x=550,y=400)
                                    
                                    lb3=Label(window,text='SEM        --',bg='skyblue',fg='red',font=('optimus',20))
                                    lb3.place(x=350,y=500)
                                    
                                    br=['1','2','3','4','5','6','7','8']
                                    cb=ttk.Combobox(window,values=br,width=92)
                                    cb.place(x=550,y=505)
                                    
                                    button=Button(window,text='NEXT',bg='yellow',fg='darkblue',font=('MEGATRON',15),command=click)
                                    button.place(x=700,y=570)
                                    

window.mainloop()